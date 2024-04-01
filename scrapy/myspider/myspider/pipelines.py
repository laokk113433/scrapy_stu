import openpyxl


class MyspiderPipeline:
    def __init__(self):
        # 创建一个 Excel 用于存取数据
        self.file = openpyxl.Workbook()
        self.sheet = None
        # 行索引
        self.row_index = 1

    def process_item(self, item, spider):
        """
        定义对文件的操作
        :param item:
        :param spider:
        :return:
        """

        school_name = item.get('高校名称及代码')

        if school_name not in self.file.sheetnames:
            self.sheet = self.file.create_sheet(school_name)
            self.update_sheet_data(item)
        else:
            self.sheet = self.file[school_name]
            self.update_sheet_data(item)

        return item

    def update_sheet_data(self, item):

        row_data = list(item.values())

        # 只爬取艺术类的专业
        if "艺术学" not in row_data:
            return ...
        else:
            # 写入数据
            if not self.sheet.cell(row=1, column=1).value:
                # 如果第一个单元格为空，则写入表头
                headers = list(item.keys())
                for index, header in enumerate(headers):
                    self.sheet.cell(row=index + 1, column=1, value=header)

            column_index = self.sheet.max_column
            for index, data in enumerate(row_data):
                if not data:
                    self.sheet.cell(row=index + 1, column=column_index + 1, value="该字段未存储")
                else:
                    self.sheet.cell(row=index + 1, column=column_index + 1, value=data)

    def close_spider(self, spider):
        # 删除默认的空Sheet
        self.file.remove(self.file['Sheet'])
        for sheet in self.file.sheetnames:
            sheet_name = self.file[sheet]
            if sheet_name.max_row == 1 and sheet_name.max_column == 1 and sheet_name.cell(row=1, column=1).value is None:
                self.file.remove(self.file[sheet])
        # 保存文件
        self.file.save('D:/素材/excel/报考院校.xlsx')
